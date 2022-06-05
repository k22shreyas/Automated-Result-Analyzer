from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import csv
import os
import pyautogui
import pandas as pd
import cv2 as cv
import pytesseract
# install all these as pip install filename, and pip install opencv-python.
  
option = webdriver.ChromeOptions()
option.add_argument("-incognito")
option.add_argument("start-maximized")
option.add_experimental_option("excludeSwitches", ['enable-automation'])
option.add_experimental_option("detach",True)

#add your chrome driver installation path
browser = webdriver.Chrome(executable_path=r'C:\Program Files (x86)\chromedriver.exe', options=option)

def fillLoginpage(usn, ite):

    browser.get("https://results.vtu.ac.in/FMEcbcs22/resultpage.php")

    #getting hold of usn and captcha input fields.

    testbox = browser.find_element_by_xpath("/html/body/div[2]/div[1]/div[2]/div/div[2]/form/div/div[2]/div[1]/div/input")
    captchabox = browser.find_element_by_xpath("/html/body/div[2]/div[1]/div[2]/div/div[2]/form/div/div[2]/div[2]/div[1]/input")

    #start with the image capta recognition procedure
    time.sleep(2)
    myScreenshot = pyautogui.screenshot(region=(970, 405, 170, 80)) #region=(horizontal pos, vertical pos, vertical ratio, horizontal ratio)
    myScreenshot.save(r'D:\web_scrap\captcha\captcha_img.png') #change according to your dir.

    os.chdir('D:\web_scrap')
    img = cv.imread('captcha\captcha_img.png',0)
    ret,thresh = cv.threshold(img,103,150,cv.THRESH_TOZERO_INV)
    #cv.imshow('Binary Threshold', thresh)
    # Using cv2.imwrite() method
    os.chdir('D:\web_scrap\captcha')
    cv.imwrite("thresh_img.png", thresh)

    time.sleep(1)
    #os.system('"wsl tesseract thresh_img.jpg result"') #tesseract is ocr function for image to text
    img2 = cv.imread('thresh_img.png',0)
    #install tesseract from https://github.com/UB-Mannheim/tesseract/wiki choose 64-bit 
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' 
    #depends on your tesseract installation path
    custom_config = r'--oem 3 --psm 6'
    captcha = pytesseract.image_to_string(img2, config=custom_config)
    
    captcha.replace(" ", "").strip()

    print("Printing solved Captcha " +captcha)
    #Exception and Error handling
    if(len(captcha)-1 != 6 ):
        return -1
    time.sleep(1)
    try:
        testbox.send_keys(usn)
        captchabox.send_keys(captcha) 
    except:
        return -1
    try:
        print(browser.current_url)
    except:
        return -1
    
    time.sleep(2)

    #sub_codes = ["18ME751", "18CS71", "18CS72","18CS744","18CS734","18CSL76","18CSP77"]
    name_loop = 0
    marks_list = []
    marks_list.append(usn)
    sub_codes = []
    for subs in range(3,sheet.max_column):
        subject = sheet.cell(row=1,column=subs)
        sub_codes.append(subject)

    try:
        for sub_code in sub_codes:
            name = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[4]").text
            internal_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[2]").int
            external_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[3]").int
            total_marks = browser.find_element_by_xpath("//*[@id='dataPrint']//*[contains(text(),'"+sub_code+"')]//following::div[4]").int

            if(name_loop == 0):
                marks_list.append(name)
            name_loop += 1
            marks_list.append(internal_marks)
            marks_list.append(external_marks)
            marks_list.append(total_marks)

    except:
        return 1
        #Error handling
    
    #fields = ["USN", "18ME751", "18CS71", "18CS72","18CS744","18CS734","18CSL76","18CSP77"]
    with open('D:\web_scrap\\result\marks.csv', 'a') as f:
    # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(marks_list)

filepath=r"D:\web_scrap\result\student_marks_list.xlsx"    #excel path
wb=load_workbook(filepath)                                                         # load into wb
sheet=wb.active                                                                    # active workbook
#store and pass current usn to function
def main():
    ite=3
    print("START")
    while ite <= sheet.max_column:
        cell_obj = sheet.cell(row=ite, column=1)
        usn = cell_obj.value
        x = fillLoginpage(usn, ite)
        print("IN MAIN FUNC") #for testing
        print(ite) #for testing
        if(x == 1):
            print(x)
            ite = ite+1
            continue
        elif(x == -1):
            print("IN ekif block")
            print(ite)
            continue
        ite = ite+1


if __name__ == "__main__":
    main()
