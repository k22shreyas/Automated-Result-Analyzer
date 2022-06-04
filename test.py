from openpyxl import load_workbook
import openpyxl
import os

wrkbk = openpyxl.load_workbook("D:\\web_scrap\\score\\result.xlsx")
sh = wrkbk.active
max_rows = sh.max_row
total = 56
row =2
sub_length = 2
while row <= max_rows:
    usn=sh.cell(row=row,column=1)
    print(usn)
    row+=1
while sub_length <= 8:
    sh.cell(row=row,column=sub_length).value = total
    sub_length+=1
wrkbk.save('D:\\web_scrap\\score\\result.xlsx')